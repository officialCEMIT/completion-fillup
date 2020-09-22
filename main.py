import openpyxl
import csv

ordinal = lambda n: "%d%s" % (n,"tsnrhtdd"[(n//10%10!=1)*(n%10<4)*n%10::4])

def main():
    wb = openpyxl.load_workbook('completion.xlsx')
    sheet = wb.worksheets[0]
    with open("class_list.csv") as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            yr = 2020 - int(row[3].split("-")[0])
            sheet.cell(row=16, column=3).value = row[3] 
            sheet.cell(row=16, column=7).value = row[0] # LAST NAME
            sheet.cell(row=16, column=9).value = row[1] # FIRST NAME
            sheet.cell(row=16, column=11).value = row[2] # MIDDLE NAME
            wb.save(f"output/{row[0]}.xlsx".lower())
if __name__ == "__main__":
    main()